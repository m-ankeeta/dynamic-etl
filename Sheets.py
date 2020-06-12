import openpyxl
import os
import xlsxwriter
import numpy as np
from datetime import date
from time import process_time

from openpyxl.utils import get_column_letter

date = date.today()
today = date.strftime("%Y%m%d")
tod = date.strftime("%d/%m/%Y")

workbook = xlsxwriter.Workbook('model_' + today + '.xlsx')

worksheet = workbook.add_worksheet()

worksheet.write('A1', 'Date')
worksheet.write('B1', 'Ticker')
worksheet.write('C1', 'Type')
worksheet.write('D1', 'Quarter')
worksheet.write('E1', 'Year')
worksheet.write('F1', 'Estimated Total Sold')
worksheet.write('G1', 'Estimated Total Sold Max')
worksheet.write('H1', 'Estimated Total Sold Min')
worksheet.write('I1', 'Forecast w/o SA Actual')
worksheet.write('J1', 'Forecast w/o SA Max')
worksheet.write('K1', 'Forecast w/o SA  Min')



t1 = process_time()

paths=[]
filename = 'C:/Users/Lenovo/Desktop/Datasheets'
arr = os.listdir(filename)
for items in arr:
    if "~$" not in items:
        paths.append(items)

fel=2
for file in paths:
    ticker = file.split(' ')[0]
    filename = r'C:/Users/Lenovo/Desktop/Datasheets/' + file
    wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    sheets = wb.get_sheet_names()
    empsheets = [sheet for sheet in sheets if "Emp" in sheet]
    regrsheets = [sheet for sheet in sheets if "Regr" in sheet]

    r = fel
    for shhet in empsheets:
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        sheetemp = wb.get_sheet_by_name(shhet)

        maxrow = sheetemp.max_row

        for rowOfCellObjects in sheetemp['E90':'I' + str(maxrow)]:
            for cellObj in rowOfCellObjects:
                if "Min" in str(cellObj.value):
                    worksheet.write('H'+str(r),sheetemp[(get_column_letter(cellObj.column+1)) + str(cellObj.row)].value)
                    worksheet.write('G'+str(r),sheetemp[(get_column_letter(cellObj.column+1)) + str(cellObj.row-1)].value)
                    worksheet.write('F'+str(r), sheetemp[(get_column_letter(cellObj.column+1)) + str(cellObj.row-2)].value)
        r = r + 1

    r = fel
    for shet in regrsheets:
        sheetregr = wb.get_sheet_by_name(shet)
        typ = shet.split('-')[-1]
        maxro = sheetregr.max_row

        for rowOfCellObjects in sheetregr['Q15':'Q' + str(maxro)]:
            for cellObj in rowOfCellObjects:
                if "Min" in str(cellObj.value):
                    worksheet.write('K'+str(r), sheetregr['R' + str(cellObj.row)].value)
                    worksheet.write('J'+str(r), sheetregr['R' + str(cellObj.row-1)].value)
                    worksheet.write('I'+str(r), sheetregr['R' + str(cellObj.row-2)].value)
                    worksheet.write('D'+str(r), sheetregr['D' + str(cellObj.row - 2)].value)
                    if "odel" not in typ:
                        worksheet.write('C'+str(r), typ)
                    else:
                        worksheet.write('C'+str(r), "Null")
                    worksheet.write('B'+str(r),ticker)
                    worksheet.write('A'+str(r), tod)
                    y = sheetregr['C' + str(cellObj.row - 2)].value
                    year = '20' + y[2:4]
                    worksheet.write("E" + str(r), int(year))
        r = r + 1

    fel = r

workbook.close()

t2 = process_time()
print("ETL Done")
print("Grand Total for all Files ",t2-t1)






